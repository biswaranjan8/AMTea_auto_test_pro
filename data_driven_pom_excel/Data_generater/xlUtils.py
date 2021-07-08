import openpyxl


def getRowCount(file_path, sheetName):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(file_path, sheetName):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file_path, sheetName, rowNo, columnNo):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNo, column=columnNo).value


def writeData(file_path, sheetName, rowNo, columnno, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNo, column=columnno).value = data
    workbook.save(file_path)
